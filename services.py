# services.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import get_connection

CATEGORIES = ["Food", "Transport", "Entertainment", "Health", "Shopping", "Utilities", "Other"]


def add_expense(amount: float, category: str, description: str, date: str):
    conn = get_connection()
    conn.execute(
        "INSERT INTO expenses (amount, category, description, date) VALUES (?, ?, ?, ?)",
        (amount, category, description, date),
    )
    conn.commit()
    conn.close()


def delete_expense(expense_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    conn.commit()
    conn.close()


def get_expenses(month: str | None = None) -> list:
    conn = get_connection()
    if month:
        rows = conn.execute(
            "SELECT * FROM expenses WHERE date LIKE ? ORDER BY date DESC",
            (f"{month}%",),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM expenses ORDER BY date DESC"
        ).fetchall()
    conn.close()
    return rows


def get_summary(month: str | None = None) -> dict:
    rows = get_expenses(month)
    total = sum(r["amount"] for r in rows)
    by_category: dict[str, float] = {}
    for row in rows:
        by_category[row["category"]] = by_category.get(row["category"], 0) + row["amount"]
    return {"total": total, "by_category": by_category}


def export_to_excel(path: str, month: str | None = None):
    rows = get_expenses(month)
    summary = get_summary(month)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transactions ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transactions"

    # styling constants
    GREEN       = "00C853"
    DARK_BG     = "0F0F23"
    SURFACE     = "16163A"
    TEXT_WHITE  = "F0F0F0"
    MUTED       = "6B6B9A"

    header_font    = Font(bold=True, color=TEXT_WHITE, size=11)
    header_fill    = PatternFill("solid", fgColor=SURFACE)
    accent_font    = Font(bold=True, color=GREEN, size=11)
    total_fill     = PatternFill("solid", fgColor=DARK_BG)
    center         = Alignment(horizontal="center", vertical="center")
    right          = Alignment(horizontal="right",  vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center")

    # Title row
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    label = f"Expenses — {month}" if month else "Expenses — All Time"
    title_cell.value     = label
    title_cell.font      = Font(bold=True, color=GREEN, size=13)
    title_cell.fill      = PatternFill("solid", fgColor=DARK_BG)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ["Date", "Category", "Description", "Amount (£)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    ws.row_dimensions[2].height = 22

    # Data rows
    for i, row in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=row["date"]).alignment       = center
        ws.cell(row=i, column=2, value=row["category"]).alignment   = left_align
        ws.cell(row=i, column=3, value=row["description"] or "—").alignment = left_align
        amount_cell = ws.cell(row=i, column=4, value=round(row["amount"], 2))
        amount_cell.alignment  = right
        amount_cell.number_format = "£#,##0.00"
        ws.row_dimensions[i].height = 20

    # Total row
    total_row = len(rows) + 3
    ws.merge_cells(f"A{total_row}:C{total_row}")
    total_label             = ws.cell(row=total_row, column=1, value="TOTAL")
    total_label.font        = Font(bold=True, color=GREEN, size=11)
    total_label.fill        = total_fill
    total_label.alignment   = right

    total_val               = ws.cell(row=total_row, column=4, value=round(summary["total"], 2))
    total_val.font          = Font(bold=True, color=GREEN, size=11)
    total_val.fill          = total_fill
    total_val.alignment     = right
    total_val.number_format = "£#,##0.00"
    ws.row_dimensions[total_row].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 14

    # ── Sheet 2: Summary ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    ws2.merge_cells("A1:B1")
    s_title           = ws2["A1"]
    s_title.value     = "Category Breakdown"
    s_title.font      = Font(bold=True, color=GREEN, size=13)
    s_title.fill      = PatternFill("solid", fgColor=DARK_BG)
    s_title.alignment = center
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(["Category", "Amount (£)", "% of Total"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
    ws2.row_dimensions[2].height = 22

    for i, (cat, total_amt) in enumerate(
        sorted(summary["by_category"].items(), key=lambda x: -x[1]), start=3
    ):
        pct = (total_amt / summary["total"] * 100) if summary["total"] else 0
        ws2.cell(row=i, column=1, value=cat).alignment              = left_align
        amt_cell = ws2.cell(row=i, column=2, value=round(total_amt, 2))
        amt_cell.alignment    = right
        amt_cell.number_format = "£#,##0.00"
        pct_cell = ws2.cell(row=i, column=3, value=round(pct, 1))
        pct_cell.alignment    = right
        pct_cell.number_format = '0.0"%"'
        ws2.row_dimensions[i].height = 20

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 14

    wb.save(path)